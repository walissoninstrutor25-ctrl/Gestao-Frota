"""Gera os dados do Controle de Escala a partir das planilhas oficiais.

Uso:
    python3 gerar_dados.py \\
        --motoristas "Escala_5x1_Motorista_Canavieiro_Safra_2027.xlsx" \\
        --lideres "Escala_6X2_Lideres_de_Turno_e_Patio_MNS.xlsx" \\
        --master "Escala_5X1_Master_Drivers_MNS_PRA.xlsx" \\
        --outdir ../data

Cada planilha precisa manter a mesma estrutura das planilhas da Safra 2026:
abas NOMES/MESTRE + uma aba por mês (MARÇO..DEZEMBRO), com os dias 1..31 nas
colunas e o status trabalha/folga indicado pela cor de preenchimento da
célula (branco = trabalha, cinza = folga) — não pelo texto da célula.

Os arquivos gerados em data/ são .js (não .json): cada um define uma
variável global `window.ALGUMA_COISA = {...}` e é incluído no index.html
via <script src="...">, em vez de ser buscado com fetch(). Isso é
proposital — fetch() de arquivo local é bloqueado pelo navegador quando a
página é aberta com duplo-clique (file://) em vez de hospedada, então o
app funciona igual hospedado ou aberto direto do disco.

Para trocar de safra/ano: atualize a constante YEAR abaixo antes de rodar.
"""
import openpyxl
import calendar
import json
import os
import re
import sys
import unicodedata

MONTHS = [
    ('MARÇO', 3), ('ABRIL', 4), ('MAIO', 5), ('JUNHO', 6), ('JULHO', 7),
    ('AGOSTO', 8), ('SETEMBRO', 9), ('OUTUBRO', 10), ('NOVEMBRO', 11), ('DEZEMBRO', 12),
]
MONTH_LABEL = {
    3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho', 7: 'Julho',
    8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro',
}
YEAR = 2026

# As únicas duas UO que aparecem nos dados recebidos (vêm da planilha do
# Master Driver). Reaproveitadas nas abas que não tinham UO nenhuma:
# a planilha de origem delas só cobre o lado MNS, então os colaboradores
# existentes são marcados MNS e o lado PRA nasce vazio, pronto pra ser
# povoado pelo próprio app.
UNIDADES_MNS_PRA = [
    {'codigo': 'MNS', 'uo': '4824', 'label': 'UO 4824 · MNS'},
    {'codigo': 'PRA', 'uo': '4823', 'label': 'UO 4823 · PRA'},
]


def is_off_fill(cell):
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    if fg.type == 'rgb':
        return fg.rgb == 'FFBFBFBF'
    if fg.type == 'theme':
        tint = fg.tint or 0
        if fg.theme == 0 and tint < -0.05:
            return True
        if fg.theme == 1 and tint > 0.05:
            return True
    return False


def find_day1_col(ws, row, search_from=2, search_to=12):
    for c in range(search_from, search_to):
        v1 = ws.cell(row=row, column=c).value
        v2 = ws.cell(row=row, column=c + 1).value
        v3 = ws.cell(row=row, column=c + 2).value
        if v1 == 1 and v2 == 2 and v3 == 3:
            return c
    return None


def norm_name(v):
    if not v:
        return None
    v = unicodedata.normalize('NFKD', str(v)).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'\s+', ' ', v).strip().upper()


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def as_matricula(v):
    v = clean(v)
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s.isdigit():
        return int(s)
    return None  # e.g. '#REF!' — broken source formula, treated as unknown


def norm_papel(papel, is_master_generic_turno=False):
    if papel is None:
        return None
    p = papel.strip()
    pu = p.upper()
    if pu in ('A', 'B', 'C'):
        return f'Turno {pu}'
    if 'FOLGUISTA' in pu:
        return 'Folguista'
    if 'APOIO' in pu:
        return p.title().replace(' A-B', ' A-B').strip()
    return p


def parse_month_sheet(ws, ndays, section_marker_fn):
    """
    section_marker_fn(text, boundary_index) -> section label.
    boundary_index is a 1-based counter of section-boundary rows seen so far
    in this sheet (title row excluded), so callers can key off row ORDER
    rather than exact text, since the source spreadsheets are not always
    spelled/labelled consistently month to month.
    """
    rows_out = []
    current_section = None
    boundary_idx = 0
    for r in range(1, ws.max_row + 1):
        a_raw = ws.cell(row=r, column=1).value
        a = clean(a_raw)
        b = clean(ws.cell(row=r, column=2).value)

        if isinstance(a, str) and a != 'COLABORADOR' and not a.upper().startswith('ESCALA') and not a.upper().startswith('ESCCALA'):
            day1_probe = find_day1_col(ws, r)
            if day1_probe is None:
                boundary_idx += 1
                current_section = section_marker_fn(a, boundary_idx)
                continue
        if a == 'COLABORADOR':
            continue
        if a is None and b is None:
            continue

        day1c = find_day1_col(ws, r)
        if day1c is None:
            continue

        nome = a if isinstance(a, str) else None
        if nome is None or nome == '#REF!':
            continue  # vacant template slot or broken reference, no real person

        papel_col = day1c - 1
        papel = clean(ws.cell(row=r, column=papel_col).value)

        matricula_raw = ws.cell(row=r, column=2).value
        matricula = as_matricula(matricula_raw)

        sched = []
        for d in range(ndays):
            cell = ws.cell(row=r, column=day1c + d)
            sched.append('O' if is_off_fill(cell) else 'W')

        rows_out.append({
            'section': current_section,
            'nome': nome,
            'matricula': matricula,
            'papel': papel,
            'schedule': ''.join(sched),
        })
    return rows_out


def parse_all_months(path, section_marker_fn, expect_title_contains=None, skip_months=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for mname, mnum in MONTHS:
        if mname not in wb.sheetnames:
            continue
        if skip_months and mname in skip_months:
            continue
        ws = wb[mname]
        ndays = calendar.monthrange(YEAR, mnum)[1]
        if expect_title_contains:
            title = clean(ws.cell(row=1, column=1).value) or ''
            if expect_title_contains.upper() not in title.upper():
                # A aba existe mas com o conteúdo errado (ex.: Dezembro do
                # Master Driver veio com uma cópia da escala de Líder de
                # Turno/Pátio). Em vez de descartar o mês inteiro, ele é
                # mantido com 0 linhas: aparece no app com os dias em
                # branco, prontos para preenchimento manual (modo de
                # edição) em vez de sumir ou mostrar dado errado.
                print(f'  ! aviso: {mname} com titulo inesperado {title!r} (esperava conter {expect_title_contains!r}) — mantido em branco para preencher manualmente', file=sys.stderr)
                result[mname] = {'mes_num': mnum, 'dias': ndays, 'linhas': []}
                continue
        rows = parse_month_sheet(ws, ndays, section_marker_fn)
        result[mname] = {'mes_num': mnum, 'dias': ndays, 'linhas': rows}
    return result


def consolidate(monthly_data, extra_lookup=None, extra_lookup_by_name=None, unidade_lookup=None, unidade_lookup_by_name=None):
    """
    monthly_data: {MONTH_KEY: {mes_num, dias, linhas:[...]}}
    Returns list of consolidated person dicts with escala per month.
    Keyed by (matricula or nome) + section, to keep distinct slots distinct.
    """
    order = [m for m, _ in MONTHS if m in monthly_data]
    people = {}
    people_order = []

    for mname in order:
        info = monthly_data[mname]
        for row in info['linhas']:
            # Keyed by (name, section) rather than matrícula: the source
            # spreadsheets' registration-number formulas are demonstrably
            # unreliable (break to '#REF!', or silently drift to a wrong
            # number in a single month — e.g. Deybd Souza Martins reads
            # 31442 in Maio and 27995 every other month) while the printed
            # name stays stable, so name is the trustworthy join key. This
            # also keeps a matrícula genuinely duplicated across two
            # sections (e.g. Advaldo Lima Lira in both Grupo 05 and Grupo
            # 07) visible as two separate people instead of one clobbering
            # the other.
            key = f"{norm_name(row['nome'])}|{row['section']}"
            if key not in people:
                people[key] = {
                    'matricula_votes': {},
                    'nome': row['nome'],
                    'grupo': row['section'],
                    'papel': row['papel'],
                    'papelNormalizado': norm_papel(row['papel']),
                    'escala': {},
                }
                people_order.append(key)
            p = people[key]
            if row['matricula']:
                p['matricula_votes'][row['matricula']] = p['matricula_votes'].get(row['matricula'], 0) + 1
            p['escala'][mname] = row['schedule']

    out = []
    for key in people_order:
        p = people[key]
        votes = p.pop('matricula_votes')
        p['matricula'] = max(votes, key=votes.get) if votes else None
        extra = None
        if extra_lookup and p['matricula'] and p['matricula'] in extra_lookup:
            extra = extra_lookup[p['matricula']]
        elif extra_lookup_by_name and norm_name(p['nome']) in (extra_lookup_by_name or {}):
            extra = extra_lookup_by_name[norm_name(p['nome'])]
        if extra:
            p.update({k: v for k, v in extra.items() if v is not None})
        unidade = None
        if unidade_lookup and p['matricula'] and p['matricula'] in unidade_lookup:
            unidade = unidade_lookup[p['matricula']]
        elif unidade_lookup_by_name and norm_name(p['nome']) in (unidade_lookup_by_name or {}):
            unidade = unidade_lookup_by_name[norm_name(p['nome'])]
        if unidade:
            p['unidade'] = unidade
        out.append(p)
    return out


def build_meses_meta(monthly_data):
    metas = []
    for mname, mnum in MONTHS:
        if mname not in monthly_data:
            continue
        metas.append({
            'chave': mname,
            'numero': mnum,
            'dias': monthly_data[mname]['dias'],
            'nome': MONTH_LABEL[mnum],
        })
    return metas


# ---------- NOMES parsers ----------

def parse_nomes_file1(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['NOMES']
    lookup = {}
    lookup_by_name = {}
    for r in range(2, ws.max_row + 1):
        nome = clean(ws.cell(row=r, column=2).value)
        mat = as_matricula(ws.cell(row=r, column=1).value)
        if mat is None and nome is None:
            continue
        lider = clean(ws.cell(row=r, column=5).value)
        telefone = clean(ws.cell(row=r, column=6).value)
        cargo = clean(ws.cell(row=r, column=4).value)
        entry = {'lider': lider, 'telefone': telefone, 'cargo': cargo}
        if mat is not None:
            lookup[mat] = entry
        if nome:
            lookup_by_name[norm_name(nome)] = entry
    return lookup, lookup_by_name


def parse_nomes_file2(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['NOMES']
    lookup = {}
    lookup_by_name = {}
    for r in range(2, ws.max_row + 1):
        nome = clean(ws.cell(row=r, column=3).value)
        mat = as_matricula(ws.cell(row=r, column=2).value)
        if mat is None and nome is None:
            continue
        # NOTE: the FUNÇÃO column in this sheet is unreliable (contradicts
        # the actual Líder de Turno/Pátio section a person is scheduled
        # under in MESTRE/monthly sheets — e.g. Diego Martins Viana is
        # tagged 'MASTER' here but is a real Líder de Pátio in the
        # schedule), so it is intentionally not surfaced in the app.
        telefone = clean(ws.cell(row=r, column=7).value)
        entry = {'telefone': telefone}
        if mat is not None:
            lookup[mat] = entry
        if nome:
            lookup_by_name[norm_name(nome)] = entry
    return lookup, lookup_by_name


def parse_nomes_file3(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['NOMES']
    unidade = {}
    unidade_by_name = {}
    for r in range(2, ws.max_row + 1):
        nome = clean(ws.cell(row=r, column=3).value)
        mat = as_matricula(ws.cell(row=r, column=2).value)
        op = clean(ws.cell(row=r, column=7).value)
        if op not in ('MNS', 'PRA'):
            continue
        if mat is not None:
            unidade[mat] = op
        if nome:
            unidade_by_name[norm_name(nome)] = op
    return unidade, unidade_by_name


def marker_grupo(a, idx):
    if a and re.match(r'^GRUPO\s*\d+', a.strip()):
        return re.sub(r'\s+', ' ', a.strip())
    return f'GRUPO desconhecido {idx}'


def marker_file2(a, idx):
    # Source sheets aren't always spelled consistently month to month
    # (e.g. Novembro mislabels the pátio block as "GRUPO 02"), so section
    # identity is taken from row ORDER: 1st boundary = Líder de Turno,
    # 2nd = Líder de Pátio, anything after (e.g. "LOGISTICA") is ignored.
    if idx == 1:
        return 'LIDER DE TURNO'
    if idx == 2:
        return 'LIDER DE PATIO'
    return 'IGNORAR'


def marker_file3(a, idx):
    # Both MNS and PRA blocks are labelled "MASTER" in the source; unit is
    # resolved later via the NOMES sheet's OP column, keyed by matrícula.
    if a.strip().upper() == 'MASTER':
        return 'MASTER'
    return 'IGNORAR'


# ---------- MESTRE parsers (equipe/equipamento structure, not month-specific) ----------

def _pessoa(mat, nome):
    return {'matricula': mat, 'nome': nome} if (mat or nome) else None


def parse_mestre_file1(path):
    """GRUPO block -> repeated 3-row 'Equipamento' sub-blocks (header row with
    turno labels + optional folguistas in cols E-M, then a matrículas row,
    then a names row). Folguistas (Turno A/B/C) only appear on the first
    equipamento header row of each grupo."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['MESTRE']
    grupos = []
    current = None
    r = 1
    while r <= ws.max_row:
        a = clean(ws.cell(row=r, column=1).value)
        if isinstance(a, str) and re.match(r'^GRUPO\s*\d+', a.strip()):
            current = {'grupo': re.sub(r'\s+', ' ', a.strip()), 'equipamentos': [], 'folguistas': {}}
            grupos.append(current)
            r += 1
            continue
        if a == 'Equipamento' and current is not None:
            for col, turno in ((5, 'A'), (8, 'B'), (11, 'C')):
                pessoa = _pessoa(as_matricula(ws.cell(row=r, column=col + 1).value), clean(ws.cell(row=r, column=col + 2).value))
                if pessoa:
                    current['folguistas'][turno] = pessoa
            numero = clean(ws.cell(row=r + 1, column=1).value)
            if numero is not None:
                turnos = {}
                for col, turno in ((2, 'A'), (3, 'B'), (4, 'C')):
                    pessoa = _pessoa(as_matricula(ws.cell(row=r + 1, column=col).value), clean(ws.cell(row=r + 2, column=col).value))
                    if pessoa:
                        turnos[turno] = pessoa
                # col A of the names row sometimes carries a status tag
                # instead of a name (e.g. 'Dedicado', 'Reserva' for spare
                # vehicles without a fixed crew) — surfaced as-is.
                status = clean(ws.cell(row=r + 2, column=1).value)
                equip = {'numero': str(numero), 'turnos': turnos}
                if isinstance(status, str):
                    equip['status'] = status
                current['equipamentos'].append(equip)
            r += 3
            continue
        r += 1
    # drop groups that carry neither equipamentos nor folguistas (empty
    # template stubs left over in the source, e.g. a trailing 'GRUPO 10'
    # with no rows filled in under it)
    return [g for g in grupos if g['equipamentos'] or g['folguistas']]


def parse_mestre_file2(path):
    """'LIDER DE TURNO' / 'LIDER DE PATIO' sections, each with Turno A/B/C
    2-row blocks (matrícula row + nome row). A Folguista rides on the first
    Turno's header row (cols C-E). An 'Apoio' pair rides in cols G/H next to
    Turno A and B's rows — the source mislabels both 'APOIO A', so position
    (not the label text) decides Apoio A vs Apoio B, matching how section
    identity elsewhere in this file is taken from row order."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['MESTRE']
    sections = []
    current = None
    apoio_idx = 0
    r = 1
    while r <= ws.max_row:
        a = clean(ws.cell(row=r, column=1).value)
        if isinstance(a, str) and a.strip().upper() in ('LIDER DE TURNO', 'LIDER DE PATIO'):
            current = {'titulo': 'Líder de Turno' if 'TURNO' in a.upper() else 'Líder de Pátio', 'turnos': {}, 'folguista': None, 'apoio': []}
            sections.append(current)
            apoio_idx = 0
            r += 1
            continue
        if a == 'Turno' and current is not None:
            folguista = _pessoa(as_matricula(ws.cell(row=r, column=4).value), clean(ws.cell(row=r, column=5).value))
            if folguista:
                current['folguista'] = folguista
            r += 1
            continue
        if a in ('A', 'B', 'C') and current is not None:
            pessoa = _pessoa(as_matricula(ws.cell(row=r, column=2).value), clean(ws.cell(row=r + 1, column=2).value))
            if pessoa:
                current['turnos'][a] = pessoa
            apoio = _pessoa(as_matricula(ws.cell(row=r, column=8).value), clean(ws.cell(row=r + 1, column=8).value))
            if apoio:
                apoio_idx += 1
                apoio['turno'] = chr(ord('A') + apoio_idx - 1)
                current['apoio'].append(apoio)
            r += 2
            continue
        r += 1
    return sections


def parse_mestre_file3(path):
    """'MASTER OP-MNS' / 'MASTER OP-PRA' / 'LIDER DE PATIO APOIO' sections,
    each with Turno A/B/C 2-row blocks (no folguista). The Apoio section is
    a placeholder in the source (matrícula/nome both literal 0) and is
    dropped once its turnos come out empty."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['MESTRE']
    sections = []
    current = None
    r = 1
    while r <= ws.max_row:
        a = clean(ws.cell(row=r, column=1).value)
        if isinstance(a, str) and a.strip().upper() in ('MASTER OP-MNS', 'MASTER OP-PRA', 'LIDER DE PATIO APOIO'):
            current = {'titulo': a.strip(), 'turnos': {}}
            sections.append(current)
            r += 1
            continue
        if a in ('A', 'B', 'C') and current is not None:
            pessoa = _pessoa(as_matricula(ws.cell(row=r, column=2).value), clean(ws.cell(row=r + 1, column=2).value))
            if pessoa:
                current['turnos'][a] = pessoa
            r += 2
            continue
        r += 1
    return [s for s in sections if s['turnos']]


def write_js_data(path, varname, data):
    """Write `data` as `window.<varname> = {...};` instead of plain JSON.

    The app loads all data through plain <script src> tags, not fetch(), so
    it works the same whether opened via file://, a local server, or a real
    host: fetch()/XHR of local files is blocked by browsers' CORS rules when
    a page is opened directly (double-clicked) instead of served over
    http(s), which is why an earlier fetch()-based version only worked when
    hosted. <script> tags aren't subject to that restriction.
    """
    with open(path, 'w', encoding='utf-8') as f:
        f.write(f'window.{varname} = ')
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write(';\n')


def build_file1(path, out_path):
    monthly = parse_all_months(path, marker_grupo)
    lookup, lookup_by_name = parse_nomes_file1(path)
    people = consolidate(monthly, extra_lookup=lookup, extra_lookup_by_name=lookup_by_name)
    # A planilha recebida só cobre o lado MNS (não tem coluna de UO como a
    # do Master Driver) — todo mundo aqui é marcado MNS, e o lado PRA
    # nasce vazio na aba, pronto pra ser preenchido pelo app.
    for p in people:
        p['unidade'] = 'MNS'
    grupos = sorted(set(p['grupo'] for p in people if p['grupo']))
    # Colaboradores are split into one small file per grupo (data/motoristas/*.js)
    # instead of one big array inline here: keeps each generated file small,
    # which matters when they need to be transferred/reviewed individually.
    outdir = os.path.dirname(out_path)
    subdir = os.path.join(outdir, 'motoristas')
    os.makedirs(subdir, exist_ok=True)
    grupo_vars = {}
    for g in grupos:
        slug = re.sub(r'\s+', '_', g.strip().lower())
        varname = f'DATA_MOTORISTAS_{slug.upper()}'
        grupo_vars[g] = varname
        write_js_data(os.path.join(subdir, f'{slug}.js'), varname, [p for p in people if p['grupo'] == g])
    data = {
        'titulo': 'Escala 5x1 — Motoristas Canavieiros',
        'tipoEscala': '5x1',
        'ano': YEAR,
        'meses': build_meses_meta(monthly),
        'unidades': UNIDADES_MNS_PRA,
        'grupos': grupos,
        'colaboradoresPorGrupoVar': grupo_vars,
        'mestre': parse_mestre_file1(path),
    }
    write_js_data(out_path, 'DATA_MOTORISTAS_META', data)
    print(f'{out_path}: {len(people)} colaboradores em {len(grupos)} arquivo(s) de grupo, grupos={grupos}')
    return data


def build_file2(path, out_turno_path, out_patio_path):
    monthly = parse_all_months(path, marker_file2)
    lookup, lookup_by_name = parse_nomes_file2(path)

    # split monthly linhas by section before consolidating
    monthly_turno = {}
    monthly_patio = {}
    for mname, info in monthly.items():
        monthly_turno[mname] = {**info, 'linhas': [r for r in info['linhas'] if r['section'] == 'LIDER DE TURNO']}
        monthly_patio[mname] = {**info, 'linhas': [r for r in info['linhas'] if r['section'] == 'LIDER DE PATIO']}

    people_turno = consolidate(monthly_turno, extra_lookup=lookup, extra_lookup_by_name=lookup_by_name)
    people_patio = consolidate(monthly_patio, extra_lookup=lookup, extra_lookup_by_name=lookup_by_name)
    # A planilha (título "ESCALA 6X2 - MNS") só cobre o lado MNS — mesmo
    # raciocínio do build_file1: marca MNS, PRA fica vazio pra preencher.
    for p in people_turno + people_patio:
        p['unidade'] = 'MNS'

    mestre_sections = parse_mestre_file2(path)
    mestre_turno = next((s for s in mestre_sections if s['titulo'] == 'Líder de Turno'), None)
    mestre_patio = next((s for s in mestre_sections if s['titulo'] == 'Líder de Pátio'), None)

    data_turno = {
        'titulo': 'Escala 6x2 — Líder de Turno',
        'tipoEscala': '6x2',
        'ano': YEAR,
        'meses': build_meses_meta(monthly_turno),
        'unidades': UNIDADES_MNS_PRA,
        'colaboradores': people_turno,
        'mestre': mestre_turno,
    }
    data_patio = {
        'titulo': 'Escala 6x2 — Líder de Pátio',
        'tipoEscala': '6x2',
        'ano': YEAR,
        'meses': build_meses_meta(monthly_patio),
        'unidades': UNIDADES_MNS_PRA,
        'colaboradores': people_patio,
        'mestre': mestre_patio,
    }
    write_js_data(out_turno_path, 'DATA_LIDERES_TURNO', data_turno)
    write_js_data(out_patio_path, 'DATA_LIDERES_PATIO', data_patio)
    print(f'{out_turno_path}: {len(people_turno)} colaboradores')
    print(f'{out_patio_path}: {len(people_patio)} colaboradores')
    return data_turno, data_patio


def build_file3(path, out_path):
    # Dezembro in this workbook was accidentally overwritten with a copy of
    # the Líder de Turno/Pátio (6x2) sheet — its title says "ESCALA 6X2 -
    # DEZEMBRO" instead of "ESCALA 5X1", and its rows are Eder/Emerson/etc,
    # not master drivers. Rather than show that wrong data (or hide the
    # month entirely), expect_title_contains makes parse_all_months keep
    # Dezembro with 0 linhas — the app shows it with blank days, ready to
    # fill in by hand until a corrected sheet is provided.
    monthly = parse_all_months(path, marker_file3, expect_title_contains='5X1')
    unidade, unidade_by_name = parse_nomes_file3(path)
    people = consolidate(monthly, unidade_lookup=unidade, unidade_lookup_by_name=unidade_by_name)
    # drop rows from the generic 'IGNORAR'/LOGISTICA block and anyone we
    # couldn't classify into MNS/PRA via the NOMES sheet
    people = [p for p in people if p['grupo'] == 'MASTER' and p.get('unidade') in ('MNS', 'PRA')]
    mestre_sections = parse_mestre_file3(path)
    mestre_by_uo = {}
    for s in mestre_sections:
        if 'MNS' in s['titulo'].upper():
            mestre_by_uo['MNS'] = s
        elif 'PRA' in s['titulo'].upper():
            mestre_by_uo['PRA'] = s
    data = {
        'titulo': 'Escala 5x1 — Master Driver — MNS & PRA',
        'tipoEscala': '5x1',
        'ano': YEAR,
        'meses': build_meses_meta(monthly),
        'unidades': UNIDADES_MNS_PRA,
        'colaboradores': people,
        'mestre': mestre_by_uo,
    }
    write_js_data(out_path, 'DATA_MASTER_DRIVER', data)
    print(f'{out_path}: {len(people)} colaboradores, unidades={[p["unidade"] for p in people]}')
    return data


def build_adm5x2(out_path):
    """Nenhuma das 3 planilhas recebidas cobre uma escala 'ADM 5x2' (5
    dias de trabalho + 2 de folga por semana, o padrão comum de horário
    administrativo) — a planilha de Motorista Canavieiro/MESTRE que foi
    enviada era só um exemplo de estrutura, não dados reais desse tipo.
    Em vez de inventar colaboradores, esta aba nasce vazia — mesma
    estrutura de mês/dias das outras (Março a Dezembro) e as mesmas duas
    UO do Master Driver (único par de UO que existe nos dados recebidos)
    — pronta para colaboradores serem adicionados pelo próprio app (modo
    de edição, botão "+ Adicionar colaborador"), que já preenche os fins
    de semana como folga automaticamente para esse tipo de escala.
    """
    meses = [{'chave': mname, 'numero': mnum, 'dias': calendar.monthrange(YEAR, mnum)[1], 'nome': MONTH_LABEL[mnum]} for mname, mnum in MONTHS]
    data = {
        'titulo': 'Escala ADM 5x2 — Administrativo',
        'tipoEscala': 'adm5x2',
        'ano': YEAR,
        'meses': meses,
        'unidades': UNIDADES_MNS_PRA,
        'colaboradores': [],
    }
    write_js_data(out_path, 'DATA_ADM5X2', data)
    print(f'{out_path}: aba em branco (2 UO), pronta para preencher no app')
    return data


if __name__ == '__main__':
    import argparse

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--motoristas', required=True, help='Escala 5x1 Motorista Canavieiro (.xlsx)')
    ap.add_argument('--lideres', required=True, help='Escala 6x2 Lideres de Turno e Patio (.xlsx)')
    ap.add_argument('--master', required=True, help='Escala 5x1 Master Drivers MNS/PRA (.xlsx)')
    ap.add_argument('--outdir', default=os.path.join(os.path.dirname(__file__), '..', 'data'))
    args = ap.parse_args()

    os.makedirs(args.outdir, exist_ok=True)
    build_file1(args.motoristas, os.path.join(args.outdir, 'motoristas.js'))
    build_file2(args.lideres, os.path.join(args.outdir, 'lideres_turno.js'), os.path.join(args.outdir, 'lideres_patio.js'))
    build_file3(args.master, os.path.join(args.outdir, 'master_driver.js'))
    build_adm5x2(os.path.join(args.outdir, 'adm5x2.js'))
